#!/usr/bin/env python3
"""Fill Trade_Desk_Checklist2.xlsx Atlas columns from the signal-engine checklist.

Modifies:
  F — wired marker
  G — Atlas mapped source
  H — gap note
  I — UI update span (how often that parameter’s value refreshes on the desk)

Columns A–E (including desk source links in E) are never changed.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CHECKLIST_PATH = ROOT / "apps" / "backend" / "src" / "app" / "domains" / "trade_desk_checklist.py"
CONSTANTS_PATH = (
    ROOT / "apps" / "backend" / "src" / "app" / "domains" / "signal_engine_constants.py"
)


def _load_module(path: Path, name: str):
    import importlib.util

    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise SystemExit(f"Cannot load module: {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _load_default_metrics() -> list[dict]:
    return list(_load_module(CHECKLIST_PATH, "trade_desk_checklist").DEFAULT_METRICS)


def _tier_ttl_ms() -> dict[str, int]:
    return dict(_load_module(CONSTANTS_PATH, "signal_engine_constants").TIER_TTL_MS)


# Columns we write (1-based): F=6 Checked, G=7 source, H=8 gap, I=9 UI span.
COL_CHECKED = 6
COL_ATLAS_SOURCE = 7
COL_GAP_NOTE = 8
COL_UI_SPAN = 9

YAHOO_FEEDS = {
    "us_futures_chg",
    "eu_futures_chg",
    "gold_chg",
    "silver_chg",
    "global_gift_nifty_chg",
    "global_nikkei_chg",
    "global_sti_chg",
    "global_hang_seng_chg",
    "global_taiwan_chg",
    "global_kospi_chg",
    "global_set_thailand_chg",
    "global_jakarta_chg",
    "global_shanghai_chg",
    "global_ftse_chg",
    "global_cac40_chg",
    "global_dax_chg",
    "global_dow_fut_chg",
    "global_sp500_fut_chg",
    "global_nasdaq_fut_chg",
    "global_dow_jones_chg",
    "global_sp500_chg",
    "global_nasdaq_chg",
    "global_asx200_chg",
    "global_bitcoin_chg",
    "global_crypto_max_abs_chg",
    "global_bond_proxy_chg",
    "europe_session_max_abs_chg",
    "dow_change_pct",
}

KITE_QUOTE_FEEDS = {
    "oi",
    "ce",
    "pe",
    "sensex_ce",
    "sensex_pe",
    "banknifty_ce",
    "banknifty_pe",
    "iv",
    "atm_volume",
    "straddle",
    "india_vix",
    "crude_ltp",
    "usd_inr",
    "index_nifty_chg",
    "index_sensex_chg",
    "index_banknifty_chg",
    "index_finnifty_chg",
    "stock_reliance_chg",
    "stock_hdfc_chg",
    "stock_infosys_chg",
    "stock_sbi_chg",
    "stock_icici_chg",
    "stock_airtel_chg",
    "nifty_points_move",
    "sensex_points_move",
    "spot_vs_open",
    "fut_basis",
    "oi_pct_chg",
    "iv_chg",
    "rsi",
    "vix_chg",
    "atm",
    "pcr",
    "max_pain",
    "spot_chg",
    "gap_pct",
    "straddle_decay_pct",
    "straddle_decay_calm_pct",
    "writer_grip_score",
}

KITE_CANDLE_FEEDS = {
    "adx",
    "prev_day_high",
    "prev_day_low",
    "pivot_point",
    "cpr_bottom",
    "cpr_top",
    "inside_first_5m_range",
    "inside_day_range",
    "spot_vs_sma20_5m",
    "first_5m_high",
    "day_high",
    "last_expiry_high",
    "last_expiry_low",
    "prev_month_expiry_high",
    "prev_month_expiry_low",
    "expiry_boundary_high",
    "expiry_boundary_low",
    "running_month_high",
    "running_month_low",
    "vwap_1m",
    "vwap_distance_pct",
    "supertrend_dir",
    "chart_1m_bar_chg_pct",
    "chart_5m_bar_chg_pct",
    "chart_60m_bar_chg_pct",
    "chart_1d_bar_chg_pct",
    "chart_1w_bar_chg_pct",
    "chart_1mo_bar_chg_pct",
    "chart_1m_post_big_move_pct",
    "chart_5m_3pm_window_pct",
}

MANUAL_CONFIG_FEEDS = {"ivp", "fii_net"}


def atlas_mapping(metric: dict) -> str:
    """One-line Atlas source for the checklist spreadsheet."""
    feed = str(metric.get("feed_key") or "").strip()
    source = str(metric.get("source") or "").strip().lower()
    rule = str(metric.get("rule") or "info")
    metric_id = str(metric.get("id") or "")

    if source == "kite_candles" or feed in KITE_CANDLE_FEEDS:
        return f"Atlas · Kite get_historical_candles · {feed or 'levels'}"

    if source == "nse" or feed == "advance_decline_ratio":
        return f"Atlas · NSE public API (slow) · {feed or 'advance_decline_ratio'}"

    if rule == "ce_pe_balance" or metric_id in {"nifty_ce_pe", "chk_005", "chk_018"}:
        ce_key = metric.get("ce_feed_key") or "ce"
        pe_key = metric.get("pe_feed_key") or "pe"
        underlying = metric.get("underlying_symbol") or "ATM"
        return f"Atlas · Kite get_quote · {ce_key} + {pe_key} ({underlying})"

    if feed in YAHOO_FEEDS:
        return f"Atlas · Yahoo Finance slow tier · {feed}"

    if feed in KITE_QUOTE_FEEDS:
        if feed == "writer_grip_score":
            return "Atlas · Kite chain OI (ATM±5) · writer_grip_score"
        if feed in {"straddle_decay_pct", "straddle_decay_calm_pct"}:
            return f"Atlas · Kite get_quote · {feed} (session straddle)"
        if feed == "pcr":
            return "Atlas · Kite chain OI (ATM±5) or manual config · pcr"
        if feed == "max_pain":
            return "Atlas · Kite chain OI (ATM±5) or manual config · max_pain"
        if feed in MANUAL_CONFIG_FEEDS:
            return f"Atlas · Kite get_quote + manual override · {feed}"
        return f"Atlas · Kite get_quote · {feed}"

    if feed in MANUAL_CONFIG_FEEDS or metric_id in {"chk_009", "fii_net"}:
        return f"Atlas · manual signal config · {feed or metric_id}"

    if rule == "before_time" or metric_id == "no_trade_after_10":
        return "Atlas · computed · ist_hour (clock)"

    if rule == "iv_pct_day_high":
        return "Atlas · Kite get_quote · iv vs session high"

    if rule == "below_prev_close":
        return "Atlas · Kite get_quote · crude_ltp vs prev close"

    if rule == "spot_below_max_pain":
        return "Atlas · Kite chain OI or manual config · max_pain"

    if rule in {"abs_lte", "lt", "gt", "lte", "gte", "between"} and feed:
        return f"Atlas · see feed · {feed}"

    if rule == "info" or not feed:
        return "Atlas · desk watch · manual (no auto feed)"

    return f"Atlas · {rule} · {feed or metric_id or 'manual'}"


def atlas_gap_note(metric: dict, mapping: str) -> str:
    """Why a row is still manual, or confirmation it is auto-wired."""
    if "manual (no auto feed)" not in mapping and "not mapped" not in mapping:
        return "Auto-wired in Atlas"

    check_no = int(metric.get("check_no") or 0)
    category = str(metric.get("category") or "")

    if category == "Trade Discipline Check":
        return "Operator self-check — no API (Column E = desk reference only)"

    if 51 <= check_no <= 56:
        return "Chart pattern review — Kite has candles; no pass/fail rule defined"

    if check_no in {21, 30, 31}:
        return "StockMojo content — no public API"

    if check_no in {22, 24, 41, 80}:
        return "News judgment — headline review, not machine-verifiable"

    if check_no in {20, 33, 34, 39}:
        return "Subjective timing / session pattern — operator confirms"

    return "Desk watch — Column E link only; no Atlas auto feed yet"


def _fmt_ms(ms: int) -> str:
    if ms < 1_000:
        return f"{ms}ms"
    if ms < 60_000:
        sec = ms / 1000
        return f"{sec:g}s"
    if ms < 3_600_000:
        mins = ms / 60_000
        return f"{mins:g} min"
    hours = ms / 3_600_000
    return f"{hours:g}h"


def atlas_ui_span(metric: dict, mapping: str, tier_ttl: dict[str, int]) -> str:
    """How often the Signal Engine / Param Chart UI refreshes this parameter’s value.

    Desk SSE paints ~8 Hz, but each metric’s underlying feed is gated by tier TTL
    (fast ≤125ms stream slot / broker ≤500ms, medium ~60s, slow ~1h).
    Manual rows do not auto-update.
    """
    if "manual (no auto feed)" in mapping or "not mapped" in mapping:
        return "n/a — no auto UI update (operator / desk watch)"

    tier = str(metric.get("tier") or "medium").strip().lower()
    ttl = int(tier_ttl.get(tier, tier_ttl.get("medium", 60_000)))
    feed = str(metric.get("feed_key") or "").strip()
    source = str(metric.get("source") or "").strip().lower()

    paint = "UI paints ~8 Hz (SSE)"

    if tier == "fast":
        detail = (
            f"value refresh ≤{_fmt_ms(ttl)} (fast); "
            "Tier-A LTP/OI ~200ms when ticker alive, REST gap ≤5s when dead"
        )
    elif tier == "medium":
        detail = f"value refresh ~{_fmt_ms(ttl)} (medium cache)"
        if source == "kite_candles" or feed in KITE_CANDLE_FEEDS:
            detail += "; candle/level recompute on medium tick"
        elif feed in MANUAL_CONFIG_FEEDS:
            detail += "; until manual config changes"
    elif tier == "slow":
        detail = f"value refresh ~{_fmt_ms(ttl)} (slow / Yahoo·NSE macro)"
    elif tier == "broker":
        detail = f"value refresh ≤{_fmt_ms(ttl)} (broker quote TTL)"
    else:
        detail = f"value refresh ~{_fmt_ms(ttl)} (tier={tier})"

    return f"{paint}; {detail}"


def metrics_by_check_no() -> dict[int, dict]:
    out: dict[int, dict] = {}
    for metric in _load_default_metrics():
        check_no = int(metric.get("check_no") or 0)
        if check_no <= 0:
            continue
        out[check_no] = metric
    return out


def update_workbook(path: Path) -> tuple[int, int, int]:
    wb = openpyxl.load_workbook(path)
    ws = wb["Trade Desk Checklist"]
    tier_ttl = _tier_ttl_ms()

    ws.cell(row=4, column=COL_ATLAS_SOURCE, value="Atlas mapped source")
    ws.cell(row=4, column=COL_GAP_NOTE, value="Atlas gap note")
    ws.cell(row=4, column=COL_UI_SPAN, value="UI update span")

    by_no = metrics_by_check_no()
    updated = 0
    wired = 0
    manual = 0
    for row in range(5, ws.max_row + 1):
        raw_no = ws.cell(row=row, column=1).value
        if raw_no is None:
            continue
        try:
            check_no = int(raw_no)
        except (TypeError, ValueError):
            continue
        metric = by_no.get(check_no)
        if metric is None:
            ws.cell(row=row, column=COL_ATLAS_SOURCE, value="Atlas · not mapped")
            ws.cell(row=row, column=COL_GAP_NOTE, value="Missing in trade_desk_checklist.py")
            ws.cell(
                row=row,
                column=COL_UI_SPAN,
                value="n/a — not mapped in Atlas",
            )
            ws.cell(row=row, column=COL_CHECKED, value=None)
            updated += 1
            manual += 1
            continue
        mapping = atlas_mapping(metric)
        gap = atlas_gap_note(metric, mapping)
        span = atlas_ui_span(metric, mapping, tier_ttl)
        ws.cell(row=row, column=COL_ATLAS_SOURCE, value=mapping)
        ws.cell(row=row, column=COL_GAP_NOTE, value=gap)
        ws.cell(row=row, column=COL_UI_SPAN, value=span)
        if "manual (no auto feed)" not in mapping and "not mapped" not in mapping:
            ws.cell(row=row, column=COL_CHECKED, value="·")
            wired += 1
        else:
            ws.cell(row=row, column=COL_CHECKED, value=None)
            manual += 1
        updated += 1

    wb.save(path)
    return updated, wired, manual


def main() -> None:
    path = Path.home() / "Downloads" / "atlasDocs" / "Trade_Desk_Checklist2.xlsx"
    if len(sys.argv) > 1:
        path = Path(sys.argv[1]).expanduser()
    if not path.exists():
        alt = Path.home() / "Downloads" / "Trade_Desk_Checklist2.xlsx"
        path = alt if alt.exists() else path
    if not path.exists():
        raise SystemExit(f"File not found: {path}")
    updated, wired, manual = update_workbook(path)
    print(
        f"Updated {updated} rows in {path}\n"
        f"  Columns touched: F (wired ·), G (Atlas source), H (gap note), I (UI update span)\n"
        f"  Columns untouched: A–E (desk source links in E preserved)\n"
        f"  Auto-wired: {wired} | Manual desk watch: {manual}"
    )


if __name__ == "__main__":
    main()
